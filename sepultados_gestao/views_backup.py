from django.contrib.admin.views.decorators import staff_member_required
from django.http import FileResponse, HttpResponseForbidden
from django.conf import settings
from django.utils.timezone import localtime, now
from django.core.management import call_command
import os
import zipfile
import tempfile

@staff_member_required
def backup_completo(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Acesso restrito ao superusuário.")

    timestamp = localtime(now()).strftime('%Y%m%d_%H%M')
    zip_filename = f"backup_completo_{timestamp}.zip"

    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, zip_filename)
    dump_path = os.path.join(temp_dir, "dump.json")

    # Gera o dump dos dados do banco
    with open(dump_path, "w", encoding="utf-8") as dump_file:
        call_command("dumpdata", "--natural-foreign", "--natural-primary", "--indent", "2", stdout=dump_file)

    # Cria o zip com dump + arquivos do MEDIA_ROOT
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Adiciona o dump do banco
        zipf.write(dump_path, "dump.json")

        # Adiciona os arquivos de mídia
        for root, _, files in os.walk(settings.MEDIA_ROOT):
            if "backups" in root.lower():
                continue
            for file in files:
                full_path = os.path.join(root, file)
                arcname = os.path.relpath(full_path, settings.MEDIA_ROOT)
                zipf.write(full_path, f"media/{arcname}")  # coloca dentro da pasta "media" no zip

    return FileResponse(open(zip_path, 'rb'), as_attachment=True, filename=zip_filename)


import os
import zipfile
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.db import models
from sepultados_gestao.models import (
    Prefeitura, Sepultado, Exumacao, Translado, ConcessaoContrato, Receita, Anexo, RegistroAuditoria
)



@staff_member_required
def backup_prefeitura_ativa(request):
    prefeitura_id = request.session.get("prefeitura_ativa_id")

    from django.contrib import messages
    from django.shortcuts import redirect

    if not prefeitura_id:
        messages.error(request, "Selecione uma prefeitura ativa antes de gerar o backup.")
        return redirect("admin:index")  # ou redirecione para a página desejada


    try:
        prefeitura_ativa = Prefeitura.objects.get(id=prefeitura_id)
    except Prefeitura.DoesNotExist:
        return HttpResponse("Prefeitura não encontrada.", status=404)

    sepultados = Sepultado.objects.filter(
        tumulo__quadra__cemiterio__prefeitura_id=prefeitura_id
    ).select_related("tumulo", "tumulo__quadra", "tumulo__quadra__cemiterio")

    contratos = ConcessaoContrato.objects.filter(prefeitura_id=prefeitura_id)
    exumacoes = Exumacao.objects.filter(prefeitura_id=prefeitura_id)
    translados = Translado.objects.filter(
        tumulo_destino__quadra__cemiterio__prefeitura_id=prefeitura_id
    ).select_related("sepultado", "tumulo_destino", "tumulo_destino__quadra", "tumulo_destino__quadra__cemiterio")

    receitas = Receita.objects.filter(prefeitura_id=prefeitura_id)
    auditorias = RegistroAuditoria.objects.filter(prefeitura_id=prefeitura_id).select_related("usuario")


    sepultados_ids = sepultados.values_list("id", flat=True)
    contratos_ids = contratos.values_list("id", flat=True)
    exumacoes_ids = exumacoes.values_list("id", flat=True)
    translados_ids = translados.values_list("id", flat=True)

    anexos = Anexo.objects.filter(
        models.Q(content_type__model="sepultado", object_id__in=sepultados_ids) |
        models.Q(content_type__model="concessaocontrato", object_id__in=contratos_ids) |
        models.Q(content_type__model="exumacao", object_id__in=exumacoes_ids) |
        models.Q(content_type__model="translado", object_id__in=translados_ids)
    ).select_related("content_type")


    # 1 - Planilha SEPULTADOS
    wb = Workbook()
    ws = wb.active
    ws.title = "Sepultados"

    headers = [
        "Número do Sepultamento", "Nome", "CPF", "Sexo", "Data Nasc.",
        "Local Nascimento", "Nacionalidade", "Cor da Pele", "Estado Civil",
        "Nome Pai", "Nome Mãe", "Profissão", "Escolaridade",
        "Data Falecimento", "Hora Falecimento", "Local Falecimento", "Causa Morte",
        "Médico", "CRM", "Idade ao Falecer",
        "Cartório Nome", "Número Registro", "Livro", "Folha", "Data Registro",
        "Túmulo", "Quadra", "Linha", "Data Sepultamento", "Observações",
        "Forma Pagamento", "Valor", "Parcelas",
        "Responsável", "CPF Resp.", "Endereço Resp.", "Telefone Resp.",
        "Exumado em", "Trasladado em"
    ]
    ws.append(headers)

    for s in sepultados:
        ws.append([
            s.numero_sepultamento or "",
            s.nome or "",
            s.cpf_sepultado or "",
            s.get_sexo_display() if s.sexo else "",
            s.data_nascimento.strftime("%d/%m/%Y") if s.data_nascimento else "",
            s.local_nascimento or "",
            s.nacionalidade or "",
            s.cor_pele or "",
            s.get_estado_civil_display() if s.estado_civil else "",
            s.nome_pai or "",
            s.nome_mae or "",
            s.profissao or "",
            s.grau_instrucao or "",
            s.data_falecimento.strftime("%d/%m/%Y") if s.data_falecimento else "",
            s.hora_falecimento.strftime("%H:%M") if s.hora_falecimento else "",
            s.local_falecimento or "",
            s.causa_morte or "",
            s.medico_responsavel or "",
            s.crm_medico or "",
            s.idade_ao_falecer or "",
            s.cartorio_nome or "",
            s.cartorio_numero_registro or "",
            s.cartorio_livro or "",
            s.cartorio_folha or "",
            s.cartorio_data_registro.strftime("%d/%m/%Y") if s.cartorio_data_registro else "",
            s.tumulo.identificador if s.tumulo else "",
            s.tumulo.quadra.codigo if s.tumulo and s.tumulo.quadra else "",
            s.tumulo.linha if s.tumulo else "",
            s.data_sepultamento.strftime("%d/%m/%Y") if s.data_sepultamento else "",
            s.observacoes or "",
            s.get_forma_pagamento_display() if s.forma_pagamento else "",
            s.valor or "",
            s.quantidade_parcelas or "",
            s.nome_responsavel or "",
            s.cpf or "",
            s.endereco or "",
            s.telefone or "",
            s.data_exumacao.strftime("%d/%m/%Y") if s.data_exumacao else "",
            s.data_translado.strftime("%d/%m/%Y") if s.data_translado else "",
        ])

    sepultados_buffer = io.BytesIO()
    wb.save(sepultados_buffer)
    sepultados_buffer.seek(0)

    # 2 - Planilha CONTRATOS
    wb_contratos = Workbook()
    ws_contratos = wb_contratos.active
    ws_contratos.title = "Contratos"

    headers_contratos = [
        "Número Contrato", "Nome", "CPF/CNPJ", "Telefone",
        "Logradouro", "Número", "Bairro", "Cidade", "Estado", "CEP",
        "Data do Contrato",
        "Túmulo", "Quadra", "Linha",
        "Forma de Pagamento", "Valor Total", "Parcelas", "Observações", "Usuário"
    ]
    ws_contratos.append(headers_contratos)

    for c in contratos:
        ws_contratos.append([
            c.numero_contrato or "",
            c.nome or "",
            c.cpf or "",
            c.telefone or "",
            c.logradouro or "",
            c.endereco_numero or "",
            c.endereco_bairro or "",
            c.endereco_cidade or "",
            c.endereco_estado or "",
            c.endereco_cep or "",
            c.data_contrato.strftime("%d/%m/%Y") if c.data_contrato else "",
            c.tumulo.identificador if c.tumulo else "",
            c.tumulo.quadra.codigo if c.tumulo and c.tumulo.quadra else "",
            c.tumulo.linha if c.tumulo else "",
            c.get_forma_pagamento_display() if c.forma_pagamento else "",
            c.valor_total or "",
            c.quantidade_parcelas or "",
            c.observacoes or "",
            str(c.usuario_registro) if c.usuario_registro else "",
        ])

    contratos_buffer = io.BytesIO()
    wb_contratos.save(contratos_buffer)
    contratos_buffer.seek(0)

    wb_exumacoes = Workbook()
    ws_exumacoes = wb_exumacoes.active
    ws_exumacoes.title = "Exumacoes"

    headers_exumacoes = [
        "Número Documento", "Data", "Motivo", "Observações",
        "Sepultado", "Túmulo", "Quadra", "Linha",
        "Nome Responsável", "CPF", "Endereço", "Telefone",
        "Forma Pagamento", "Valor", "Parcelas"
    ]
    ws_exumacoes.append(headers_exumacoes)

    for e in exumacoes:
        ws_exumacoes.append([
            e.numero_documento or "",
            e.data.strftime("%d/%m/%Y") if e.data else "",
            e.motivo or "",
            e.observacoes or "",
            e.sepultado.nome if e.sepultado else "",
            e.tumulo.identificador if e.tumulo else "",
            e.tumulo.quadra.codigo if e.tumulo and e.tumulo.quadra else "",
            e.tumulo.linha if e.tumulo else "",
            e.nome_responsavel or "",
            e.cpf or "",
            e.endereco or "",
            e.telefone or "",
            e.get_forma_pagamento_display() if e.forma_pagamento else "",
            e.valor or "",
            e.quantidade_parcelas or "",
        ])

    exumacoes_buffer = io.BytesIO()
    wb_exumacoes.save(exumacoes_buffer)
    exumacoes_buffer.seek(0)

    # Queryset
    translados = Translado.objects.filter(
        sepultado__tumulo__quadra__cemiterio__prefeitura_id=prefeitura_id
    ).select_related(
        "sepultado", "tumulo_destino", "tumulo_destino__quadra", "tumulo_destino__quadra__cemiterio"
    )

    # Planilha Translados
    wb_translados = Workbook()
    ws_translados = wb_translados.active
    ws_translados.title = "Translados"

    headers_translados = [
        "Número Documento", "Data", "Motivo", "Observações", "Sepultado",
        "Destino", "Túmulo Destino", "Quadra", "Linha",
        "Cemitério", "Endereço do Cemitério",
        "Nome Responsável", "CPF", "Endereço", "Telefone",
        "Forma Pagamento", "Valor", "Parcelas"
    ]
    ws_translados.append(headers_translados)

    for t in translados:
        sepultado_str = str(t.sepultado) if t.sepultado else ""

        tumulo_destino = t.tumulo_destino
        quadra = tumulo_destino.quadra if tumulo_destino else None
        cemiterio = quadra.cemiterio if quadra else None

        cemiterio_nome = (
            cemiterio.nome if cemiterio else t.cemiterio_nome or ""
        )
        cemiterio_endereco = getattr(cemiterio, "endereco", "") if cemiterio else t.cemiterio_endereco or ""


        ws_translados.append([
            t.numero_documento or "",
            t.data.strftime("%d/%m/%Y") if t.data else "",
            t.motivo or "",
            t.observacoes or "",
            sepultado_str,
            t.get_destino_display() if t.destino else "",
            tumulo_destino.identificador if tumulo_destino else "",
            quadra.codigo if quadra else "",
            tumulo_destino.linha if tumulo_destino else "",
            cemiterio_nome,
            cemiterio_endereco,
            t.nome_responsavel or "",
            t.cpf or "",
            t.endereco or "",
            t.telefone or "",
            t.get_forma_pagamento_display() if t.forma_pagamento else "",
            t.valor or "",
            t.quantidade_parcelas or "",
        ])

    translados_buffer = io.BytesIO()
    wb_translados.save(translados_buffer)
    translados_buffer.seek(0)

    wb_receitas = Workbook()
    ws_receitas = wb_receitas.active
    ws_receitas.title = "Receitas"

    headers_receitas = [
        "Número", "Nome", "CPF/CNPJ", "Descrição",
        "Valor Total", "Desconto", "Valor Pago", "Valor em Aberto",
        "Data Vencimento", "Data Pagamento",
        "Status", "Multa", "Juros", "Mora Diária",
        "Contrato", "Exumação", "Translado", "Sepultado"
    ]
    ws_receitas.append(headers_receitas)

    for r in receitas:
        ws_receitas.append([
            r.numero_documento or "",
            r.nome or "",
            r.cpf or "",
            r.descricao or "",
            r.valor_total or "",
            r.desconto or "",
            r.valor_pago or "",
            r.valor_em_aberto or "",
            r.data_vencimento.strftime("%d/%m/%Y") if r.data_vencimento else "",
            r.data_pagamento.strftime("%d/%m/%Y") if r.data_pagamento else "",
            r.get_status_display() if r.status else "",
            r.multa or "",
            r.juros or "",
            r.mora_diaria or "",
            str(r.contrato.numero_contrato) if r.contrato else "",
            str(r.exumacao.numero_documento) if r.exumacao else "",
            str(r.translado.numero_documento) if r.translado else "",
            str(r.sepultado.nome) if r.sepultado else "",
        ])

    receitas_buffer = io.BytesIO()
    wb_receitas.save(receitas_buffer)
    receitas_buffer.seek(0)

    wb_auditoria = Workbook()
    ws_auditoria = wb_auditoria.active
    ws_auditoria.title = "Auditoria"

    headers_auditoria = [
        "Ação", "Usuário", "Modelo", "ID do Objeto", "Representação", "Data e Hora"
    ]
    ws_auditoria.append(headers_auditoria)

    for reg in auditorias:
        ws_auditoria.append([
            reg.get_acao_display() if reg.acao else "",
            str(reg.usuario) if reg.usuario else "",
            reg.modelo or "",
            reg.objeto_id or "",
            reg.representacao or "",
            reg.data_hora.strftime("%d/%m/%Y %H:%M") if reg.data_hora else "",
        ])

    auditoria_buffer = io.BytesIO()
    wb_auditoria.save(auditoria_buffer)
    auditoria_buffer.seek(0)


    # Criar ZIP e adicionar tudo
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        zip_file.writestr('sepultados.xlsx', sepultados_buffer.getvalue())
        zip_file.writestr('contratos.xlsx', contratos_buffer.getvalue())
        zip_file.writestr("exumacoes.xlsx", exumacoes_buffer.read())
        zip_file.writestr("translados.xlsx", translados_buffer.getvalue())
        zip_file.writestr("receitas.xlsx", receitas_buffer.read())
        zip_file.writestr("auditoria.xlsx", auditoria_buffer.getvalue())


        # Adicionar os arquivos de mídia
        for anexo in anexos:
            if anexo.arquivo:
                caminho = os.path.join(settings.MEDIA_ROOT, anexo.arquivo.name)
                if os.path.exists(caminho):
                    zip_file.write(caminho, arcname=os.path.join('midia', os.path.basename(anexo.arquivo.name)))

    from django.utils.timezone import localtime, now
    from django.utils.text import slugify

    nome_prefeitura = slugify(prefeitura_ativa.nome)
    agora = localtime(now())
    filename = f"backup_{nome_prefeitura}_{prefeitura_id}_{agora.strftime('%Y%m%d_%H%M%S')}.zip"

    return HttpResponse(
        zip_buffer.getvalue(),
        content_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



from django.shortcuts import render

@staff_member_required
def backup_sistema(request):
    return render(request, "admin/backup_sistema.html")
